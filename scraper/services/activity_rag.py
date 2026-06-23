"""
activity_rag.py
===============================================================================
RAG-based Activity Reconciliation Engine

Place this file at:
    my_scraper/scraper/services/activity_rag.py

Usage:
    from scraper.services.activity_rag import ActivityRAG

    rag = ActivityRAG("path/to/Consolidated List of Activities.xlsx")
    report = rag.reconcile(scraped_df, jurisdiction="Meydan")
    rag.save()   # overwrites master file with updates
===============================================================================
"""

from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity




def _normalize(text: str) -> str:
    """Lowercase, strip accents, collapse whitespace, remove punctuation."""
    if not isinstance(text, str):
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _clean_code(code) -> str:
    """Strip trailing .0 and whitespace from a code value."""
    return re.sub(r"\.0$", "", str(code).strip()).strip()




class _TFIDF:
    """
    Lightweight in-memory TF-IDF index.
    Supports incremental add_document() after initial build.
    """

    def __init__(self):
        self.vocab: dict[str, int] = {}
        self.matrix: Optional[np.ndarray] = None
        self._raw_docs: list[list[str]] = []
        self._dirty = True

    def fit(self, texts: list[str]):
        self._raw_docs = [_normalize(t).split() for t in texts]
        all_tokens = {tok for doc in self._raw_docs for tok in doc}
        self.vocab = {t: i for i, t in enumerate(sorted(all_tokens))}
        self._dirty = True
        self._rebuild()

    def _rebuild(self):
        n_docs  = len(self._raw_docs)
        n_terms = len(self.vocab)
        if n_docs == 0 or n_terms == 0:
            self.matrix = np.zeros((0, 0), dtype=np.float32)
            self._dirty = False
            return

        tf = np.zeros((n_docs, n_terms), dtype=np.float32)
        for d, tokens in enumerate(self._raw_docs):
            for tok in tokens:
                if tok in self.vocab:
                    tf[d, self.vocab[tok]] += 1
            if tokens:
                tf[d] /= len(tokens)

        df_counts = (tf > 0).sum(axis=0)
        idf = np.log((n_docs + 1) / (df_counts + 1)) + 1
        mat = tf * idf

        norms = np.linalg.norm(mat, axis=1, keepdims=True)
        norms[norms == 0] = 1
        self.matrix = (mat / norms).astype(np.float32)
        self._dirty = False

    def add_document(self, text: str):
        """Append one document; index will be rebuilt on next query."""
        self._raw_docs.append(_normalize(text).split())
        for tok in self._raw_docs[-1]:
            if tok not in self.vocab:
                self.vocab[tok] = len(self.vocab)
        self._dirty = True

    def query(self, text: str, top_k: int = 5) -> list[tuple[float, int]]:
        """Return [(similarity, doc_index), ...] sorted by descending similarity."""
        if self._dirty:
            self._rebuild()
        if self.matrix is None or self.matrix.shape[0] == 0:
            return []

        tokens = _normalize(text).split()
        vec = np.zeros(len(self.vocab), dtype=np.float32)
        for tok in tokens:
            if tok in self.vocab:
                vec[self.vocab[tok]] += 1
        norm = np.linalg.norm(vec)
        if norm == 0:
            return []
        vec /= norm

        sims = self.matrix @ vec
        top  = np.argsort(sims)[::-1][:top_k]
        return [(float(sims[i]), int(i)) for i in top]




@dataclass
class ActivityChange:
    activity_code:    str
    jurisdiction:     str
    scraped_name:     str
    master_name:      str
    change_type:      str          
    assigned_code:    str   = ""
    similarity_score: float = 0.0
    nearest_neighbor: str   = ""


@dataclass
class ReconciliationReport:
    jurisdiction:   str
    updated:        list[ActivityChange] = field(default_factory=list)
    new_activities: list[ActivityChange] = field(default_factory=list)
    unchanged:      list[ActivityChange] = field(default_factory=list)
    skipped:        list[ActivityChange] = field(default_factory=list)

    def summary(self) -> str:
        lines = [
            f"\n{'='*58}",
            f"  Reconciliation -- {self.jurisdiction}",
            f"{'='*58}",
            f"  OK  Unchanged     : {len(self.unchanged)}",
            f"  >>  Name Updated  : {len(self.updated)}",
            f"  +   New Activities: {len(self.new_activities)}",
            f"  --  Skipped       : {len(self.skipped)}",
            f"{'-'*58}",
        ]
        if self.updated:
            lines.append("\n  NAME UPDATES:")
            for c in self.updated:
                lines.append(f"    [{c.activity_code}] '{c.master_name}' -> '{c.scraped_name}'")
        if self.new_activities:
            lines.append("\n  NEW ACTIVITIES:")
            for c in self.new_activities:
                nb = (f"  (nearest: '{c.nearest_neighbor}', sim={c.similarity_score:.2f})"
                      if c.nearest_neighbor else "  (fresh code)")
                lines.append(f"    [{c.assigned_code}] '{c.scraped_name}'{nb}")
        return "\n".join(lines)




class ActivityRAG:
    """
    RAG engine over the master activity Excel file.

    Responsibilities
    ----------------
    1. Load the "Final" sheet into memory.
    2. Build a TF-IDF index over all activity names.
    3. reconcile(scraped_df, jurisdiction):
         A. (code, jurisdiction) exists  -> update name if different
         B. code exists, jurisdiction new -> insert row for this jurisdiction
         C. code unknown / empty          -> assign code, insert row
    4. save(path) -> write updated master back, preserving all other sheets.

    Similarity thresholds
    ---------------------
    NAME_UPDATE_THRESHOLD : min similarity to accept a scraped name as a valid
                            correction for the master  (default 0.30)
    CODE_ASSIGN_THRESHOLD : min similarity to reuse an existing code for a
                            brand-new activity         (default 0.20)
    """

    NAME_UPDATE_THRESHOLD = 0.30
    CODE_ASSIGN_THRESHOLD = 0.20

    # master sheet column names (lowercased)
    _NAME  = "activity name"
    _CODE  = "activity code"
    _JURIS = "jurisdiction"
    _DIV   = "division"
    _GRP   = "group"
    _CLS   = "class"
    _ISIC  = "isic description"
    _DESC  = "activity description"

    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self._load()
        self._build_index()

    def _load(self):
        self.df = pd.read_excel(self.xlsx_path, sheet_name="Final", dtype=str)
        self.df.columns = self.df.columns.str.strip().str.lower()

        self.df[self._CODE]  = self.df[self._CODE].fillna("").apply(_clean_code)
        self.df[self._NAME]  = self.df[self._NAME].fillna("").str.strip()
        self.df[self._JURIS] = self.df[self._JURIS].fillna("").str.strip()

        int_codes = [int(c) for c in self.df[self._CODE] if re.fullmatch(r"\d+", c)]
        self._max_code: int = max(int_codes) if int_codes else 100000

        print(f"[RAG] Loaded {len(self.df)} rows | "
              f"{self.df[self._JURIS].nunique()} jurisdictions | "
              f"max code = {self._max_code}")



    def _build_index(self):
        self._cj_idx: dict[tuple, int]     = {}       # (code, juris) -> row index
        self._c_idx:  dict[str, list[int]] = defaultdict(list)  # code -> [row indices]

        for i, row in self.df.iterrows():
            code  = row[self._CODE]
            juris = row[self._JURIS]
            if code:
                self._cj_idx[(code, juris)] = i
                self._c_idx[code].append(i)

        print(f"[RAG] Building TF-IDF index for {len(self.df)} activity names...")
        self._vectorizer = TfidfVectorizer(
            analyzer='word',
            token_pattern=r"\b\w+\b",
            lowercase=False,
            norm='l2',
            max_df=0.95,
            min_df=1,
        )
        self._tfidf_matrix = self._vectorizer.fit_transform(
            self.df[self._NAME].fillna("").astype(str).tolist()
        )
        print(f"[RAG] TF-IDF matrix shape: {self._tfidf_matrix.shape}")

        print(f"[RAG] Index ready: {len(self._cj_idx)} (code, jurisdiction) pairs")

 

    def find_similar(self, name: str, top_k: int = 5) -> list[tuple[float, int]]:
        """Return [(similarity, row_index), ...] for the closest activity names."""
        if not hasattr(self, '_tfidf_matrix') or self._tfidf_matrix is None or self._tfidf_matrix.shape[0] == 0:
            return []

        query_vec = self._vectorizer.transform([name])
        sims = cosine_similarity(self._tfidf_matrix, query_vec).flatten()
        top = np.argsort(sims)[::-1][:top_k]
        return [(float(sims[i]), int(i)) for i in top if sims[i] > 0]

    def _assign_code(self, name: str) -> tuple[str, float, str]:
        """
        Pick a code for a new activity.
        Reuses nearest neighbour's code if similarity >= CODE_ASSIGN_THRESHOLD,
        otherwise allocates max_code + 1.
        Returns (code, similarity, nearest_name).
        """
        results = self.find_similar(name, top_k=1)
        if results:
            sim, idx = results[0]
            nearest_name = self.df.at[idx, self._NAME]
            nearest_code = self.df.at[idx, self._CODE]
            if sim >= self.CODE_ASSIGN_THRESHOLD and nearest_code:
                return nearest_code, round(sim, 4), nearest_name

        self._max_code += 1
        return str(self._max_code), 0.0, ""



    def reconcile(
        self,
        scraped_df: pd.DataFrame,
        jurisdiction: str,
        code_col: str = "activity code",
        name_col: str = "activity name",
    ) -> ReconciliationReport:
        """
        Compare scraped_df against master for the given jurisdiction only.
        Mutations are staged in self.df; call save() to persist.
        """
        report = ReconciliationReport(jurisdiction=jurisdiction)

        scraped_df = scraped_df.copy()
        scraped_df.columns = scraped_df.columns.str.strip().str.lower()

        # Accept "code" as alias
        if "code" in scraped_df.columns and code_col not in scraped_df.columns:
            scraped_df = scraped_df.rename(columns={"code": code_col})

        if code_col in scraped_df.columns:
            scraped_df[code_col] = scraped_df[code_col].fillna("").apply(_clean_code)
        else:
            scraped_df[code_col] = ""

        if name_col in scraped_df.columns:
            scraped_df[name_col] = scraped_df[name_col].fillna("").astype(str).str.strip()
        else:
            scraped_df[name_col] = ""

        new_rows: list[dict] = []

        for _, srow in scraped_df.iterrows():
            code         = srow.get(code_col, "")
            scraped_name = srow.get(name_col, "")

            if not scraped_name:
                report.skipped.append(ActivityChange(
                    activity_code=code, jurisdiction=jurisdiction,
                    scraped_name="", master_name="", change_type="skipped",
                ))
                continue

            cj_key = (code, jurisdiction)

         
            if code and cj_key in self._cj_idx:
                idx         = self._cj_idx[cj_key]
                master_name = self.df.at[idx, self._NAME]

                if _normalize(scraped_name) == _normalize(master_name):
                    report.unchanged.append(ActivityChange(
                        activity_code=code, jurisdiction=jurisdiction,
                        scraped_name=scraped_name, master_name=master_name,
                        change_type="unchanged",
                    ))
                else:
                    sim_results = self.find_similar(scraped_name, top_k=1)
                    sim         = sim_results[0][0] if sim_results else 0.0

                    if sim >= self.NAME_UPDATE_THRESHOLD or master_name == "":
                        self.df.at[idx, self._NAME] = scraped_name
                        # Keep TF-IDF doc in sync
                        self._tfidf._raw_docs[idx] = _normalize(scraped_name).split()
                        self._tfidf._dirty = True

                        report.updated.append(ActivityChange(
                            activity_code=code, jurisdiction=jurisdiction,
                            scraped_name=scraped_name, master_name=master_name,
                            change_type="updated",
                            similarity_score=round(sim, 4),
                        ))
                    else:
                        report.skipped.append(ActivityChange(
                            activity_code=code, jurisdiction=jurisdiction,
                            scraped_name=scraped_name, master_name=master_name,
                            change_type="skipped",
                            similarity_score=round(sim, 4),
                        ))

          
            elif code and code in self._c_idx:
                ref_row = self.df.iloc[self._c_idx[code][0]]
                new_rows.append({
                    self._NAME:  scraped_name,
                    self._CODE:  code,
                    self._DIV:   ref_row.get(self._DIV,  ""),
                    self._GRP:   ref_row.get(self._GRP,   ""),
                    self._CLS:   ref_row.get(self._CLS,   ""),
                    self._ISIC:  ref_row.get(self._ISIC,  ""),
                    self._DESC:  "",
                    self._JURIS: jurisdiction,
                })
                report.new_activities.append(ActivityChange(
                    activity_code=code, jurisdiction=jurisdiction,
                    scraped_name=scraped_name,
                    master_name=ref_row.get(self._NAME, ""),
                    change_type="new_activity", assigned_code=code,
                ))

            
            else:
                assigned_code, sim, nearest = self._assign_code(scraped_name)
                new_rows.append({
                    self._NAME:  scraped_name,
                    self._CODE:  assigned_code,
                    self._DIV:   "",
                    self._GRP:   "",
                    self._CLS:   "",
                    self._ISIC:  "",
                    self._DESC:  "",
                    self._JURIS: jurisdiction,
                })
                report.new_activities.append(ActivityChange(
                    activity_code=code or "(none)",
                    jurisdiction=jurisdiction,
                    scraped_name=scraped_name, master_name="",
                    change_type="new_activity", assigned_code=assigned_code,
                    similarity_score=sim, nearest_neighbor=nearest,
                ))

        # Append all new rows, then rebuild index
        if new_rows:
            self.df = pd.concat(
                [self.df, pd.DataFrame(new_rows)],
                ignore_index=True,
            )
            self._build_index()

        print(report.summary())
        return report

    

    def save(self, output_path: Optional[str] = None) -> str:
        """
        Write the updated master DataFrame back to xlsx.
        All other sheets (Consolidated Sheet, ISIC, Sheet3) are preserved unchanged.
        """
        out = output_path or self.xlsx_path

        wb = load_workbook(self.xlsx_path)
        if "Final" in wb.sheetnames:
            del wb["Final"]
        ws = wb.create_sheet("Final")

        cols = [
            self._NAME, self._CODE,
            self._DIV, self._GRP, self._CLS,
            self._ISIC, self._DESC, self._JURIS,
        ]

        thin      = Side(style="thin", color="D9D9D9")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        hdr_fill  = PatternFill("solid", start_color="1F3864")
        data_font = Font(name="Arial", size=10)
        center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)

        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c_idx, value=col)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.border = border; cell.alignment = center
        ws.row_dimensions[1].height = 28

        for r_idx, (_, row) in enumerate(self.df.iterrows(), start=2):
            for c_idx, col in enumerate(cols, 1):
                val  = row.get(col, "") if col in self.df.columns else ""
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font   = data_font
                cell.border = border
                cell.alignment = center if c_idx in (2, 3, 4, 5, 8) else left

        for i, w in enumerate([45, 14, 10, 10, 10, 50, 30, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        try:
            wb.save(out)
            print(f"[RAG] Saved -> {out}  ({len(self.df)} rows)")
            return out
        except PermissionError as e:
            print(f"[RAG] Permission denied: {e}. Please close the Excel file if it's open and try again.")
            return None
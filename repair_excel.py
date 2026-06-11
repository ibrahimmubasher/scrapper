"""
Run this once to repair the corrupted xlsx file.
It reads what it can from the broken file and
writes a clean new copy.

Usage:
    python repair_excel.py
"""
import os
import shutil
import zipfile
import pandas as pd
from openpyxl import load_workbook


XLSX_PATH = os.path.join(
    os.getcwd(),
    "scraper",
    "data",
    "Consolidated List of Activities.xlsx"
)


def try_read_sheet(path, sheet_name):
    """Try multiple engines to read a sheet."""

    for engine in ["openpyxl", "xlrd", "calamine"]:
        try:
            df = pd.read_excel(
                path,
                sheet_name=sheet_name,
                dtype=str,
                engine=engine
            )
            print(
                f"  ✅ Read '{sheet_name}' "
                f"via {engine} "
                f"({len(df)} rows)"
            )
            return df
        except Exception as e:
            print(f"  ⚠ {engine} failed: {e}")

    return None


def repair():

    print(f"\nTarget file:\n  {XLSX_PATH}\n")

    if not os.path.exists(XLSX_PATH):
        print("❌ File not found at that path.")
        print("Edit XLSX_PATH in this script to match your file location.")
        return

    # ── Backup the corrupted file first ──
    backup_path = XLSX_PATH + ".corrupted_backup"

    if not os.path.exists(backup_path):
        shutil.copy2(XLSX_PATH, backup_path)
        print(f"✅ Backup saved:\n  {backup_path}\n")
    else:
        print(f"✅ Backup already exists.\n")

    # ── Try to read all sheets ──
    print("Reading sheets from corrupted file...")

    # Get sheet names — try zipfile directly
    sheet_names = []

    try:
        with zipfile.ZipFile(XLSX_PATH, "r") as z:
            names = z.namelist()
            # xl/worksheets/sheet1.xml etc
            sheets_raw = [
                n for n in names
                if n.startswith("xl/worksheets/sheet")
                and n.endswith(".xml")
            ]
            print(f"  ZIP entries found: {len(names)}")

        # Try openpyxl with keep_vba / read_only
        try:
            wb = load_workbook(
                XLSX_PATH,
                read_only=True,
                data_only=True
            )
            sheet_names = wb.sheetnames
            wb.close()
            print(f"  Sheets: {sheet_names}")
        except Exception as e:
            print(f"  Could not read sheet names: {e}")
            # Fallback: assume standard sheet names
            sheet_names = ["Final", "ISIC"]
            print(f"  Assuming sheets: {sheet_names}")

    except Exception as e:
        print(f"  ZIP read failed: {e}")
        sheet_names = ["Final", "ISIC"]

    # ── Read each sheet ──
    recovered = {}

    for sheet in sheet_names:
        print(f"\nReading sheet: '{sheet}'")
        df = try_read_sheet(XLSX_PATH, sheet)
        if df is not None:
            recovered[sheet] = df
        else:
            print(f"  ❌ Could not recover sheet '{sheet}'")

    if not recovered:
        print(
            "\n❌ Could not recover any sheets.\n"
            "Please restore from a backup copy manually."
        )
        return

    # ── Write clean new file ──
    clean_path = XLSX_PATH  # overwrite in place

    print(f"\nWriting clean file...")

    try:
        with pd.ExcelWriter(
            clean_path,
            engine="openpyxl"
        ) as writer:
            for sheet_name, df in recovered.items():
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False
                )
                print(
                    f"  ✅ Written: '{sheet_name}' "
                    f"({len(df)} rows)"
                )

        print(f"\n✅ Repaired file saved:\n  {clean_path}")

    except Exception as e:
        print(f"\n❌ Write failed: {e}")
        print("Restoring backup...")
        shutil.copy2(backup_path, XLSX_PATH)
        print("Backup restored.")
        return

    # ── Verify the repaired file ──
    print("\nVerifying repaired file...")

    try:
        for sheet_name in recovered:
            df_check = pd.read_excel(
                clean_path,
                sheet_name=sheet_name,
                dtype=str
            )
            print(
                f"  ✅ '{sheet_name}': "
                f"{len(df_check)} rows — OK"
            )
        print("\n✅ Repair complete. Run your scraper now.")

    except Exception as e:
        print(f"\n❌ Verification failed: {e}")
        print("Restoring backup...")
        shutil.copy2(backup_path, XLSX_PATH)
        print("Backup restored. Please fix manually.")


if __name__ == "__main__":
    repair()
